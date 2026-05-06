import streamlit as st
import json, re, io
from groq import Groq
from tavily import TavilyClient
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, PageBreak, HRFlowable
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────
st.set_page_config(page_title="AI Roadmap Generator", page_icon="🗺️")

st.title("🗺️ AI Learning Roadmap Generator")

# ─────────────────────────────────────────────────────────
# SIDEBAR (API KEYS)
# ─────────────────────────────────────────────────────────
with st.sidebar:
    st.header("API Keys")
    groq_key = st.text_input("Groq API Key", type="password")
    tavily_key = st.text_input("Tavily API Key (optional)", type="password")

# ─────────────────────────────────────────────────────────
# INPUTS
# ─────────────────────────────────────────────────────────
topic = st.text_input("What do you want to learn?")
level = st.selectbox("Level", ["beginner", "intermediate", "advanced"])
days = st.slider("Days", 3, 30, 14)
hours = st.slider("Hours per day", 1, 8, 2)

# ─────────────────────────────────────────────────────────
# FUNCTIONS
# ─────────────────────────────────────────────────────────

def fetch_resources(topic, level, api_key):
    if not api_key:
        return []
    try:
        client = TavilyClient(api_key=api_key)
        results = client.search(
            query=f"best resources to learn {topic} for {level}",
            max_results=5
        )
        return [{"title": r["title"], "url": r["url"]} for r in results["results"]]
    except:
        return []


def generate_roadmap(topic, days, hours, level, resources, api_key):
    client = Groq(api_key=api_key)

    prompt = f"""
Create a {days}-day roadmap for learning {topic}.
Level: {level}, Hours/day: {hours}

Return JSON with:
title, overview, totalDays, hoursPerDay, level,
phases, days, resources
"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.7
    )

    raw = response.choices[0].message.content.strip()
    raw = re.sub(r'```json|```', '', raw)

    return json.loads(raw)


def generate_pdf(roadmap):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(roadmap["title"], styles["Title"]))
    story.append(Spacer(1, 10))

    for day in roadmap["days"]:
        story.append(Paragraph(f"Day {day['day']}: {day['title']}", styles["Heading2"]))
        story.append(Paragraph(day["description"], styles["Normal"]))
        story.append(Spacer(1, 8))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_excel(roadmap):
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    ws.append(["Day", "Title", "Description"])

    for d in roadmap["days"]:
        ws.append([d["day"], d["title"], d["description"]])

    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────────────────
# MAIN BUTTON
# ─────────────────────────────────────────────────────────
if st.button("Generate Roadmap"):

    if not topic:
        st.error("Enter a topic")
    elif not groq_key:
        st.error("Enter Groq API key")
    else:
        with st.spinner("Fetching resources..."):
            resources = fetch_resources(topic, level, tavily_key)

        with st.spinner("Generating roadmap..."):
            roadmap = generate_roadmap(
                topic, days, hours, level, resources, groq_key
            )

        st.success("Roadmap generated!")

        # SHOW OUTPUT
        st.subheader(roadmap["title"])
        st.write(roadmap["overview"])

        for day in roadmap["days"]:
            with st.expander(f"Day {day['day']} - {day['title']}"):
                st.write(day["description"])

        # DOWNLOADS
        pdf = generate_pdf(roadmap)
        excel = generate_excel(roadmap)

        st.download_button("Download PDF", pdf, "roadmap.pdf")
        st.download_button("Download Excel", excel, "roadmap.xlsx")
