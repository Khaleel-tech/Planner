import streamlit as st
import json, re
from groq import Groq
from tavily import TavilyClient
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, PageBreak, HRFlowable
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AI Learning Roadmap Agent",
    page_icon="🗺️",
    layout="wide"
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,300&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Hide default Streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 4rem; max-width: 1100px; }

/* Hero header */
.hero-wrap {
    background: linear-gradient(135deg, #0a0a0a 0%, #1a1a2e 50%, #0a0a0a 100%);
    border-radius: 20px;
    padding: 3rem 2.5rem;
    margin-bottom: 2rem;
    border: 1px solid #222;
    position: relative;
    overflow: hidden;
}
.hero-wrap::before {
    content: '';
    position: absolute;
    top: -50%; left: -50%;
    width: 200%; height: 200%;
    background: radial-gradient(circle at 30% 50%, rgba(99,102,241,0.08) 0%, transparent 50%),
                radial-gradient(circle at 70% 50%, rgba(16,185,129,0.06) 0%, transparent 50%);
    pointer-events: none;
}
.hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 2.6rem;
    font-weight: 800;
    color: #ffffff;
    margin: 0 0 0.5rem 0;
    line-height: 1.1;
    letter-spacing: -0.02em;
}
.hero-title span { color: #818cf8; }
.hero-sub {
    font-size: 1rem;
    color: #888;
    margin: 0;
    font-weight: 300;
    letter-spacing: 0.01em;
}
.hero-badges {
    display: flex;
    gap: 8px;
    margin-top: 1.5rem;
    flex-wrap: wrap;
}
.badge {
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 20px;
    padding: 4px 14px;
    font-size: 12px;
    color: #aaa;
    font-family: 'DM Sans', sans-serif;
}

/* Section labels */
.section-label {
    font-family: 'Syne', sans-serif;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #818cf8;
    margin-bottom: 0.75rem;
}

/* Card container */
.card {
    background: #0f0f0f;
    border: 1px solid #1f1f1f;
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}

/* Stat cards */
.stat-row {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin: 1.5rem 0;
}
.stat-card {
    background: #0f0f0f;
    border: 1px solid #1f1f1f;
    border-radius: 14px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.stat-num {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    color: #ffffff;
    line-height: 1;
    margin-bottom: 4px;
}
.stat-lbl {
    font-size: 11px;
    color: #555;
    text-transform: uppercase;
    letter-spacing: 0.08em;
}

/* Phase pills */
.phase-strip {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    margin: 1rem 0 1.5rem;
}
.phase-pill {
    border-radius: 30px;
    padding: 6px 16px;
    font-size: 12px;
    font-weight: 500;
    font-family: 'DM Sans', sans-serif;
}

/* Day card */
.day-card {
    background: #0f0f0f;
    border: 1px solid #1a1a1a;
    border-radius: 12px;
    padding: 1rem 1.25rem;
    margin-bottom: 8px;
    transition: border-color 0.2s;
}
.day-card:hover { border-color: #2a2a2a; }
.day-card.milestone { border-color: #2d2a1a; background: #0f0e09; }
.day-num {
    font-size: 10px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #444;
    margin-bottom: 4px;
}
.day-title {
    font-family: 'Syne', sans-serif;
    font-size: 15px;
    font-weight: 600;
    color: #e0e0e0;
    margin-bottom: 6px;
}
.day-desc { font-size: 13px; color: #666; line-height: 1.6; margin-bottom: 10px; }
.task-list { display: flex; flex-wrap: wrap; gap: 6px; }
.task-pill {
    background: #161616;
    border: 1px solid #222;
    border-radius: 20px;
    padding: 3px 12px;
    font-size: 11px;
    color: #555;
}
.milestone-badge {
    display: inline-block;
    background: #2d2a0a;
    color: #d4a820;
    border-radius: 20px;
    padding: 2px 10px;
    font-size: 10px;
    font-weight: 600;
    margin-left: 8px;
    letter-spacing: 0.05em;
}

/* Overview box */
.overview-box {
    background: #0a0a14;
    border: 1px solid #1a1a2e;
    border-radius: 14px;
    padding: 1.25rem 1.5rem;
    margin: 1rem 0 1.5rem;
    font-size: 14px;
    color: #888;
    line-height: 1.7;
}

/* Resource links */
.resource-item {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 10px 14px;
    background: #0f0f0f;
    border: 1px solid #1a1a1a;
    border-radius: 10px;
    margin-bottom: 6px;
    text-decoration: none;
}
.resource-dot { width: 6px; height: 6px; border-radius: 50%; background: #818cf8; flex-shrink: 0; }
.resource-title { font-size: 13px; color: #c0c0c0; }
.resource-type { margin-left: auto; font-size: 10px; color: #444; text-transform: uppercase; letter-spacing: 0.05em; }

/* Download section */
.download-section {
    background: linear-gradient(135deg, #0d0d1a, #0a0a0a);
    border: 1px solid #1a1a2e;
    border-radius: 16px;
    padding: 1.5rem;
    margin-top: 2rem;
}
.download-title {
    font-family: 'Syne', sans-serif;
    font-size: 16px;
    font-weight: 700;
    color: #fff;
    margin-bottom: 0.5rem;
}
.download-sub { font-size: 12px; color: #555; margin-bottom: 1.25rem; }

/* Streamlit button overrides */
.stButton > button {
    background: #818cf8 !important;
    color: #000 !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    padding: 0.6rem 1.5rem !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background: #a5b4fc !important;
    transform: translateY(-1px) !important;
}
.stDownloadButton > button {
    background: #0f0f0f !important;
    color: #e0e0e0 !important;
    border: 1px solid #2a2a2a !important;
    border-radius: 10px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    border-color: #818cf8 !important;
    color: #818cf8 !important;
}

/* Input overrides */
.stTextInput > div > div > input,
.stSelectbox > div > div,
.stSlider {
    background: #0f0f0f !important;
    color: #e0e0e0 !important;
    border-color: #1f1f1f !important;
}
label { color: #888 !important; font-size: 12px !important; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: #080808 !important;
    border-right: 1px solid #151515 !important;
}
[data-testid="stSidebar"] * { color: #aaa !important; }

/* Success / error */
.stSuccess { background: #0a1a0a !important; border-color: #1a3a1a !important; color: #4ade80 !important; }
.stAlert { border-radius: 10px !important; }

/* Spinner */
.stSpinner > div { border-top-color: #818cf8 !important; }

/* Expander — hide in favor of custom cards */
.streamlit-expanderHeader { display: none !important; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
PHASE_FILLS = ["B5D4F4","C0DD97","FAC775","F4C0D1","CECBF6"]
PHASE_TEXT  = ["042C53","173404","412402","4B1528","26215C"]
PHASE_COLORS_CSS = [
    ("rgba(181,212,244,0.15)", "#B5D4F4", "#042C53"),
    ("rgba(192,221,151,0.15)", "#C0DD97", "#173404"),
    ("rgba(250,199,117,0.15)", "#FAC775", "#412402"),
    ("rgba(244,192,209,0.15)", "#F4C0D1", "#4B1528"),
    ("rgba(206,203,246,0.15)", "#CECBF6", "#26215C"),
]
PDF_FILLS = [colors.HexColor(f"#{c}") for c in PHASE_FILLS]
PDF_TEXT  = [colors.HexColor(f"#{c}") for c in PHASE_TEXT]

# ── Session state init ────────────────────────────────────────────────────────
for key in ("roadmap", "pdf_buf", "xl_buf"):
    if key not in st.session_state:
        st.session_state[key] = None

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🔑 API Keys")
    groq_key   = st.text_input("Groq API Key",            type="password", placeholder="gsk_...")
    tavily_key = st.text_input("Tavily API Key (optional)",type="password", placeholder="tvly-...")

    # Fall back to Streamlit secrets if inputs are empty
    if not groq_key:
        groq_key   = st.secrets.get("GROQ_API_KEY", "")
    if not tavily_key:
        tavily_key = st.secrets.get("TAVILY_API_KEY", "")

    st.markdown("---")
    st.markdown("**Stack**")
    st.caption("🤖  LLM: Llama 3.3 70B via Groq")
    st.caption("🔍  Search: Tavily web search")
    st.caption("📄  PDF: ReportLab Platypus")
    st.caption("📊  Excel: openpyxl (3 sheets)")
    st.markdown("---")
    st.caption("Keys are used only for this session and never stored.")

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-wrap">
  <p class="hero-title">AI Learning<br><span>Roadmap Agent</span></p>
  <p class="hero-sub">Enter a topic + number of days → get a personalized day-by-day study plan</p>
  <div class="hero-badges">
    <span class="badge">⚡ Powered by Groq</span>
    <span class="badge">🔍 Tavily Web Search</span>
    <span class="badge">📄 PDF Export</span>
    <span class="badge">📊 Excel Tracker</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Input form ────────────────────────────────────────────────────────────────
c1, c2, c3 = st.columns([3, 1.2, 1.2])
with c1:
    st.markdown('<p class="section-label">Topic</p>', unsafe_allow_html=True)
    topic = st.text_input("topic", label_visibility="collapsed",
                          placeholder="e.g. Machine Learning, React, Data Structures, SQL...")
with c2:
    st.markdown('<p class="section-label">Days</p>', unsafe_allow_html=True)
    days = st.slider("days", label_visibility="collapsed", min_value=3, max_value=90, value=14)
with c3:
    st.markdown('<p class="section-label">Hours / day</p>', unsafe_allow_html=True)
    hours = st.slider("hours", label_visibility="collapsed", min_value=1, max_value=8, value=2)

c4, c5 = st.columns([2, 4])
with c4:
    st.markdown('<p class="section-label">Current level</p>', unsafe_allow_html=True)
    level = st.selectbox("level", label_visibility="collapsed",
                         options=["complete beginner","some basics","intermediate","advanced"])

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
generate_btn = st.button("✦ Generate Roadmap", use_container_width=False)

st.markdown("---")

# ── Core functions ────────────────────────────────────────────────────────────
def fetch_resources(topic, level, api_key):
    if not api_key:
        return []
    try:
        client  = TavilyClient(api_key=api_key)
        results = client.search(
            query=f"best resources to learn {topic} for {level}",
            search_depth="basic", max_results=5
        )
        return [{"title": r["title"], "url": r["url"], "type": "web"}
                for r in results.get("results", [])]
    except:
        return []

def generate_roadmap(topic, days, hours, level, resources, api_key):
    client = Groq(api_key=api_key)
    resource_context = ""
    if resources:
        lines = [f"{i+1}. {r['title']}: {r['url']}" for i, r in enumerate(resources)]
        resource_context = "\n\nReal web resources — include these URLs:\n" + "\n".join(lines)

    prompt = f"""You are an expert learning roadmap designer.
Create a detailed {days}-day learning roadmap for someone who wants to learn "{topic}".
The student is a {level} and can study {hours} hours per day.{resource_context}

Return ONLY valid JSON — no markdown, no explanation, no code fences.
{{
  "title": "Learning Roadmap: {topic}",
  "overview": "2-3 sentence summary of the full learning journey",
  "totalDays": {days},
  "hoursPerDay": {hours},
  "level": "{level}",
  "phases": [{{"name":"phase name","days":"Day 1-N","goal":"what student achieves"}}],
  "days": [{{
    "day": 1, "phase": 0,
    "title": "short engaging title",
    "description": "specific and actionable, 2-3 sentences",
    "tasks": ["task 1","task 2","task 3"],
    "milestone": false
  }}],
  "resources": [{{"title":"name","url":"https://...","type":"free|paid|book|video"}}]
}}
Rules:
- Exactly {days} day entries
- 3-5 phases: foundation → core → practice → projects → mastery
- milestone: true on every 7th day
- phase is 0-based index into phases array
- include provided URLs in resources
- every day must have specific topic names, not vague instructions
"""
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.7, max_tokens=6000
    )
    raw = response.choices[0].message.content.strip()
    raw = re.sub(r'^```json\s*','',raw)
    raw = re.sub(r'^```\s*','',raw)
    raw = re.sub(r'\s*```$','',raw)
    return json.loads(raw)

def generate_pdf(roadmap):
    buf    = io.BytesIO()
    PAGE_W = A4[0] - 4*cm
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
             rightMargin=2*cm, leftMargin=2*cm,
             topMargin=2*cm,   bottomMargin=2*cm)

    S = lambda name, **kw: ParagraphStyle(name, parent=styles["Normal"], **kw)
    S_TITLE  = S("t",  parent=styles["Title"],   fontSize=22, spaceAfter=6,   textColor=colors.HexColor("#111111"))
    S_META   = S("m",  fontSize=10, spaceAfter=12, textColor=colors.HexColor("#666666"))
    S_SEC    = S("s",  parent=styles["Heading2"], fontSize=13, spaceBefore=18, spaceAfter=8, textColor=colors.HexColor("#222222"))
    S_DTITLE = S("dt", fontSize=12, fontName="Helvetica-Bold", textColor=colors.HexColor("#111111"))
    S_DDESC  = S("dd", fontSize=10, leading=14, textColor=colors.HexColor("#444444"))
    S_TASK   = S("tk", fontSize=9,  leading=13, leftIndent=8, textColor=colors.HexColor("#555555"))
    S_RES    = S("r",  fontSize=10, leading=14, textColor=colors.HexColor("#185FA5"))
    S_FOOT   = S("f",  fontSize=8,  alignment=TA_CENTER, textColor=colors.HexColor("#AAAAAA"))

    story = []
    story.append(Paragraph(roadmap["title"], S_TITLE))
    story.append(Paragraph(
        f"{roadmap['totalDays']} days  ·  {roadmap['hoursPerDay']} hrs/day  ·  "
        f"{roadmap['level']}  ·  {roadmap['hoursPerDay']*roadmap['totalDays']}h total", S_META))

    ov_box = Table([[Paragraph(roadmap["overview"],
                    S("ov", fontSize=11, leading=16, textColor=colors.HexColor("#333333")))
                  ]], colWidths=[PAGE_W])
    ov_box.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F8F8F8")),
        ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E0E0E0")),
        ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("LEFTPADDING",(0,0),(-1,-1),12),("RIGHTPADDING",(0,0),(-1,-1),12),
    ]))
    story.append(ov_box)
    story.append(Spacer(1,14))

    story.append(Paragraph("Learning phases", S_SEC))
    pdata = [["Phase","Days","Goal"]]
    for ph in roadmap["phases"]:
        pdata.append([ph["name"], ph["days"], ph.get("goal","")])
    pt = Table(pdata, colWidths=[PAGE_W*0.25, PAGE_W*0.15, PAGE_W*0.60])
    pts = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#222222")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),10),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F7F7F7")]),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#E0E0E0")),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
    ])
    for i in range(1, len(pdata)):
        idx = (i-1) % len(PDF_FILLS)
        pts.add("BACKGROUND",(0,i),(0,i),PDF_FILLS[idx])
        pts.add("TEXTCOLOR",(0,i),(0,i),PDF_TEXT[idx])
        pts.add("FONTNAME",(0,i),(0,i),"Helvetica-Bold")
    pt.setStyle(pts)
    story.append(pt)
    story.append(Spacer(1,16))

    story.append(Paragraph("Day-by-day plan", S_SEC))
    for day in roadmap["days"]:
        idx   = (day.get("phase") or 0) % len(PDF_FILLS)
        label = f"Day {day['day']}" + (" ★" if day.get("milestone") else "")
        tasks = "     ".join(f"• {t}" for t in day.get("tasks",[]))
        lp = Paragraph(label, S("dl", fontSize=9, fontName="Helvetica-Bold", textColor=PDF_TEXT[idx]))
        lc = Table([[lp]], style=TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),PDF_FILLS[idx]),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ]))
        inner = Table(
            [[lc, Paragraph(day["title"], S_DTITLE)],
             ["",  Paragraph(day["description"], S_DDESC)],
             ["",  Paragraph(tasks, S_TASK)]],
            colWidths=[PAGE_W*0.14, PAGE_W*0.86]
        )
        inner.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("LEFTPADDING",(1,0),(1,-1),10),("LEFTPADDING",(0,0),(0,-1),0),
        ]))
        outer = Table([[inner]], colWidths=[PAGE_W])
        outer.setStyle(TableStyle([
            ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E0E0E0")),
            ("BACKGROUND",(0,0),(-1,-1),colors.white),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ]))
        story.append(outer)
        story.append(Spacer(1,5))

    story.append(PageBreak())
    story.append(Paragraph("Recommended resources", S_SEC))
    for r in roadmap.get("resources",[]):
        story.append(Paragraph(
            f"→  <a href='{r['url']}'><u>{r['title']}</u></a>"
            f"  <font color='#888888' size='9'>({r.get('type','resource')})</font>", S_RES))
        story.append(Spacer(1,4))

    story.append(Spacer(1,30))
    story.append(HRFlowable(width=PAGE_W, thickness=0.5, color=colors.HexColor("#E0E0E0")))
    story.append(Spacer(1,8))
    story.append(Paragraph("Generated by AI Learning Roadmap Agent", S_FOOT))

    doc.build(story)
    buf.seek(0)
    return buf

def generate_excel(roadmap):
    buf    = io.BytesIO()
    wb     = Workbook()
    phases = roadmap.get("phases", [])
    total  = roadmap["totalDays"]

    THIN  = Side(style="thin", color="E0E0E0")
    BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    AC    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    HF    = Font(bold=True, color="FFFFFF", size=10)
    HFill = PatternFill("solid", fgColor="222222")

    # Sheet 1 — Timetable
    ws1 = wb.active
    ws1.title = "Timetable"
    ws1.merge_cells("A1:I1")
    ws1["A1"] = roadmap["title"]
    ws1["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill      = PatternFill("solid", fgColor="111111")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:I2")
    ws1["A2"] = f"{total} days  |  {roadmap['hoursPerDay']} hrs/day  |  {roadmap['level']}  |  {roadmap['hoursPerDay']*total}h total"
    ws1["A2"].font      = Font(size=9, color="666666", italic=True)
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 22

    headers    = ["Day","Phase","Title","Description","Tasks","Milestone","Est. Hours","Completed ✓","Notes"]
    col_widths = [6, 18, 28, 48, 42, 11, 12, 14, 28]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.font = HF; c.fill = HFill; c.alignment = AC; c.border = BDR
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[3].height = 20

    for day in roadmap["days"]:
        row    = day["day"] + 3
        ph_idx = (day.get("phase") or 0) % len(PHASE_FILLS)
        pFill  = PatternFill("solid", fgColor=PHASE_FILLS[ph_idx])
        pFont  = Font(size=9, bold=True, color=PHASE_TEXT[ph_idx])
        pname  = phases[day.get("phase",0)]["name"] if phases else ""
        tasks  = "\n".join(f"• {t}" for t in day.get("tasks",[]))
        mstone = "★ Milestone" if day.get("milestone") else ""
        vals   = [day["day"], pname, day["title"], day["description"],
                  tasks, mstone, roadmap["hoursPerDay"], "", ""]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.border = BDR; c.alignment = AL; c.font = Font(size=9)
            if col in (1,2): c.fill = pFill; c.font = pFont
            if col in (1,6,7,8): c.alignment = AC
            if mstone and col not in (1,2):
                c.fill = PatternFill("solid", fgColor="FFFBF0")
        ws1.row_dimensions[row].height = max(30, 14*len(day.get("tasks",[])))

    ws1.freeze_panes = "A4"
    ws1.auto_filter.ref = f"A3:I{total+3}"

    # Sheet 2 — Progress Dashboard
    ws2 = wb.create_sheet("Progress Dashboard")
    ws2["A1"] = "Progress Dashboard"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 40
    for col, lbl in enumerate(["Metric","Value"], 1):
        c = ws2.cell(row=3, column=col, value=lbl)
        c.font = HF; c.fill = HFill; c.alignment = AC

    metrics = [
        ("Total days",        total),
        ("Hours per day",     roadmap["hoursPerDay"]),
        ("Total study hours", roadmap["hoursPerDay"]*total),
        ("Phases",            len(phases)),
        ("Days completed",    f'=COUNTIF(Timetable!H4:H{total+3},"✓")'),
        ("Days remaining",    f"={total}-B8"),
        ("% Complete",        "=ROUND(B8/B4*100,1)"),
    ]
    for i, (lbl, val) in enumerate(metrics, 4):
        ws2.cell(row=i,column=1,value=lbl).font  = Font(size=10)
        ws2.cell(row=i,column=2,value=val).font  = Font(size=10,bold=True)
        ws2.cell(row=i,column=1).alignment = Alignment(horizontal="left")
        ws2.cell(row=i,column=2).alignment = Alignment(horizontal="center")
        af = PatternFill("solid", fgColor="F7F7F7" if i%2==0 else "FFFFFF")
        ws2.cell(row=i,column=1).fill = af
        ws2.cell(row=i,column=2).fill = af
        for col in (1,2): ws2.cell(row=i,column=col).border = BDR

    ws2["A12"] = "Phase breakdown"
    ws2["A12"].font = Font(bold=True, size=11)
    for col, lbl in enumerate(["Phase","Days","Goal"],1):
        c = ws2.cell(row=13,column=col,value=lbl)
        c.font=HF; c.fill=HFill; c.alignment=AC
    for i, ph in enumerate(phases):
        row=14+i; idx=i%len(PHASE_FILLS)
        for col, val in enumerate([ph["name"],ph["days"],ph.get("goal","")],1):
            c = ws2.cell(row=row,column=col,value=val)
            c.font      = Font(size=9,bold=(col<=2),color=PHASE_TEXT[idx])
            c.fill      = PatternFill("solid",fgColor=PHASE_FILLS[idx])
            c.alignment = AL; c.border = BDR

    # Sheet 3 — Resources
    ws3 = wb.create_sheet("Resources")
    ws3["A1"] = "Learning Resources"
    ws3["A1"].font = Font(bold=True,size=14)
    ws3.column_dimensions["A"].width=35
    ws3.column_dimensions["B"].width=55
    ws3.column_dimensions["C"].width=12
    for col, lbl in enumerate(["Title","URL","Type"],1):
        c = ws3.cell(row=3,column=col,value=lbl)
        c.font=HF; c.fill=HFill; c.alignment=AC
    for i, r in enumerate(roadmap.get("resources",[]),4):
        af = PatternFill("solid",fgColor="F7F7F7" if i%2==0 else "FFFFFF")
        for col, val in enumerate([r["title"],r["url"],r.get("type","")],1):
            c = ws3.cell(row=i,column=col,value=val)
            c.font      = Font(size=10,color="185FA5" if col==2 else "111111")
            c.fill      = af; c.alignment=AL; c.border=BDR

    wb.save(buf)
    buf.seek(0)
    return buf

# ── Generate on button click ──────────────────────────────────────────────────
if generate_btn:
    if not topic:
        st.error("Please enter a topic to learn.")
    elif not groq_key:
        st.error("Groq API key not found. Add it in the sidebar or Streamlit Secrets.")
    else:
        with st.spinner("🔍  Searching web resources via Tavily..."):
            resources = fetch_resources(topic, level, tavily_key)

        with st.spinner("🤖  Generating roadmap with Llama 3.3 70B..."):
            try:
                st.session_state.roadmap  = generate_roadmap(topic, days, hours, level, resources, groq_key)
            except Exception as e:
                st.error(f"Generation failed: {e}")
                st.stop()

        with st.spinner("📄  Building PDF..."):
            st.session_state.pdf_buf = generate_pdf(st.session_state.roadmap)

        with st.spinner("📊  Building Excel tracker..."):
            st.session_state.xl_buf  = generate_excel(st.session_state.roadmap)

# ── Display results (persists via session_state) ──────────────────────────────
if st.session_state.roadmap:
    roadmap = st.session_state.roadmap
    phases  = roadmap.get("phases", [])

    # Success banner
    st.markdown(f"""
    <div style="background:#0a140a;border:1px solid #1a3a1a;border-radius:12px;
                padding:12px 18px;margin-bottom:1.5rem;display:flex;align-items:center;gap:10px;">
      <span style="color:#4ade80;font-size:16px;">✓</span>
      <span style="color:#4ade80;font-size:14px;font-weight:500;">
        Roadmap generated — {len(roadmap['days'])} days across {len(phases)} phases
      </span>
    </div>
    """, unsafe_allow_html=True)

    # Title + overview
    st.markdown(f"""
    <h2 style="font-family:'Syne',sans-serif;font-size:1.8rem;font-weight:700;
               color:#fff;margin:0 0 0.25rem;">{roadmap['title']}</h2>
    <div class="overview-box">{roadmap['overview']}</div>
    """, unsafe_allow_html=True)

    # Stats row
    total_tasks = sum(len(d.get("tasks",[])) for d in roadmap["days"])
    milestones  = sum(1 for d in roadmap["days"] if d.get("milestone"))
    st.markdown(f"""
    <div class="stat-row">
      <div class="stat-card">
        <div class="stat-num">{roadmap['totalDays']}</div>
        <div class="stat-lbl">Days</div>
      </div>
      <div class="stat-card">
        <div class="stat-num">{roadmap['hoursPerDay'] * roadmap['totalDays']}h</div>
        <div class="stat-lbl">Total study</div>
      </div>
      <div class="stat-card">
        <div class="stat-num">{total_tasks}</div>
        <div class="stat-lbl">Tasks</div>
      </div>
      <div class="stat-card">
        <div class="stat-num">{milestones}</div>
        <div class="stat-lbl">Milestones</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Phase pills
    phase_pills = ""
    for i, ph in enumerate(phases):
        idx = i % len(PHASE_COLORS_CSS)
        bg, border, txt = PHASE_COLORS_CSS[idx]
        phase_pills += f'<span class="phase-pill" style="background:{bg};border:1px solid {border}40;color:{border};">{ph["name"]} · {ph["days"]}</span>'
    st.markdown(f'<div class="phase-strip">{phase_pills}</div>', unsafe_allow_html=True)

    # Day cards — two columns
    st.markdown('<p class="section-label">Day-by-day plan</p>', unsafe_allow_html=True)
    left_col, right_col = st.columns(2)
    mid = (len(roadmap["days"]) + 1) // 2

    for col_idx, (col, day_slice) in enumerate([(left_col, roadmap["days"][:mid]),
                                                 (right_col, roadmap["days"][mid:])]):
        with col:
            for day in day_slice:
                ph_idx   = (day.get("phase") or 0) % len(PHASE_COLORS_CSS)
                _, border, _ = PHASE_COLORS_CSS[ph_idx]
                milestone_html = '<span class="milestone-badge">★ MILESTONE</span>' if day.get("milestone") else ""
                tasks_html = "".join(f'<span class="task-pill">{t}</span>' for t in day.get("tasks",[]))
                card_class = "day-card milestone" if day.get("milestone") else "day-card"
                st.markdown(f"""
                <div class="{card_class}" style="border-left:3px solid {border}40;">
                  <div class="day-num">Day {day['day']}{milestone_html}</div>
                  <div class="day-title">{day['title']}</div>
                  <div class="day-desc">{day['description']}</div>
                  <div class="task-list">{tasks_html}</div>
                </div>
                """, unsafe_allow_html=True)

    # Resources
    if roadmap.get("resources"):
        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
        st.markdown('<p class="section-label">Recommended resources</p>', unsafe_allow_html=True)
        res_cols = st.columns(2)
        for i, r in enumerate(roadmap["resources"]):
            with res_cols[i % 2]:
                st.markdown(f"""
                <a href="{r['url']}" target="_blank" class="resource-item">
                  <div class="resource-dot"></div>
                  <span class="resource-title">{r['title']}</span>
                  <span class="resource-type">{r.get('type','')}</span>
                </a>
                """, unsafe_allow_html=True)

    # Download section
    st.markdown("""
    <div class="download-section">
      <div class="download-title">Export your roadmap</div>
      <div class="download-sub">PDF for reading · Excel for tracking your daily progress</div>
    </div>
    """, unsafe_allow_html=True)

    safe_name = roadmap['title'].replace(' ','_').replace(':','')
    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            label="📄  Download PDF Roadmap",
            data=st.session_state.pdf_buf,
            file_name=f"{safe_name}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    with d2:
        st.download_button(
            label="📊  Download Excel Tracker",
            data=st.session_state.xl_buf,
            file_name=f"{safe_name}_tracker.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
